"""Spreadsheet Parser - Parse CSV/XLSX into typed chunks with header-repeated chunking."""

import csv
import io
import logging
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class SheetChunk:
    """A chunk from a spreadsheet sheet."""
    text: str
    sheet_name: str
    chunk_type: str  # "record" | "document"


class SpreadsheetParser:
    """Parse spreadsheets (CSV/XLSX) into typed chunks."""

    MAX_ROWS_PER_CHUNK = 30

    def parse(self, content: bytes, mime_type: str) -> list[SheetChunk]:
        """Parse spreadsheet content into chunks.

        Args:
            content: Raw file bytes
            mime_type: MIME type of the file

        Returns:
            List of SheetChunk with text, sheet_name, and chunk_type
        """
        if mime_type == "text/csv":
            return self._parse_csv(content)
        elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return self._parse_xlsx(content)
        else:
            # xls or unknown — plain text fallback
            return self._fallback_text(content)

    def _parse_csv(self, content: bytes) -> list[SheetChunk]:
        """Parse CSV as a single sheet."""
        text = self._decode(content)
        reader = csv.reader(io.StringIO(text))
        rows = [[cell for cell in row] for row in reader]
        return self._process_sheet(rows, "Sheet1")

    def _parse_xlsx(self, content: bytes) -> list[SheetChunk]:
        """Parse XLSX with per-sheet processing."""
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        all_chunks = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            all_chunks.extend(self._process_sheet(rows, sheet_name))

        wb.close()
        return all_chunks

    def _fallback_text(self, content: bytes) -> list[SheetChunk]:
        """Fallback: treat as plain text."""
        text = self._decode(content)
        if text.strip():
            return [SheetChunk(text=text, sheet_name="Sheet1", chunk_type="document")]
        return []

    def _process_sheet(self, rows: list[list[str]], sheet_name: str) -> list[SheetChunk]:
        """Process a single sheet's rows into chunks."""
        # Strip trailing empty rows
        while rows and all(not cell.strip() for cell in rows[-1]):
            rows.pop()

        if not rows:
            return []

        # Split into regions by 2+ consecutive blank rows
        regions = self._split_regions(rows)
        chunks = []

        for region in regions:
            sheet_type = self._detect_sheet_type(region)
            if sheet_type == "record":
                header = region[0]
                data_rows = region[1:]
                chunks.extend(self._chunk_records(header, data_rows, sheet_name))
            else:
                chunks.extend(self._chunk_document(region, sheet_name))

        return chunks

    def _split_regions(self, rows: list[list[str]]) -> list[list[list[str]]]:
        """Split rows into regions separated by 2+ consecutive blank rows."""
        regions = []
        current_region = []
        blank_count = 0

        for row in rows:
            is_blank = all(not cell.strip() for cell in row)
            if is_blank:
                blank_count += 1
                if blank_count >= 2 and current_region:
                    regions.append(current_region)
                    current_region = []
            else:
                if blank_count == 1 and current_region:
                    # Single blank row — keep in region
                    current_region.append([""] * len(row))
                blank_count = 0
                current_region.append(row)

        if current_region:
            regions.append(current_region)

        return regions

    def _detect_sheet_type(self, rows: list[list[str]]) -> str:
        """Detect whether rows represent record (tabular) or document data.

        Record signals (need 3+ to classify as record):
        - 90%+ rows have the same column count
        - Average cell length < 80 chars
        - 10+ rows
        - First row has short texts (header candidate)
        """
        if len(rows) < 2:
            return "document"

        signals = 0

        # Signal 1: consistent column count (90%+)
        col_counts = [len(row) for row in rows]
        if col_counts:
            most_common = max(set(col_counts), key=col_counts.count)
            consistent_ratio = col_counts.count(most_common) / len(col_counts)
            if consistent_ratio >= 0.9 and most_common > 1:
                signals += 1

        # Signal 2: short cells on average
        all_cells = [cell for row in rows for cell in row if cell.strip()]
        if all_cells:
            avg_len = sum(len(c) for c in all_cells) / len(all_cells)
            if avg_len < 80:
                signals += 1

        # Signal 3: 10+ rows
        if len(rows) >= 10:
            signals += 1

        # Signal 4: first row has short texts (header candidate)
        first_row_cells = [c for c in rows[0] if c.strip()]
        if first_row_cells:
            max_header_len = max(len(c) for c in first_row_cells)
            if max_header_len < 50:
                signals += 1

        return "record" if signals >= 3 else "document"

    def _chunk_records(
        self,
        header: list[str],
        data_rows: list[list[str]],
        sheet_name: str,
    ) -> list[SheetChunk]:
        """Chunk record-type data with header repeated per chunk."""
        if not data_rows:
            # Header only — still produce a chunk
            header_line = " | ".join(cell.strip() for cell in header if cell.strip())
            text = f"[Sheet: {sheet_name}]\n{header_line}"
            return [SheetChunk(text=text, sheet_name=sheet_name, chunk_type="record")]

        header_line = " | ".join(cell.strip() for cell in header)
        chunks = []

        for i in range(0, len(data_rows), self.MAX_ROWS_PER_CHUNK):
            batch = data_rows[i:i + self.MAX_ROWS_PER_CHUNK]
            lines = [f"[Sheet: {sheet_name}]", header_line]
            for row in batch:
                # Pad/trim row to match header length
                padded = row[:len(header)]
                while len(padded) < len(header):
                    padded.append("")
                lines.append(" | ".join(cell.strip() for cell in padded))

            text = "\n".join(lines)
            chunks.append(SheetChunk(text=text, sheet_name=sheet_name, chunk_type="record"))

        return chunks

    def _chunk_document(
        self,
        rows: list[list[str]],
        sheet_name: str,
    ) -> list[SheetChunk]:
        """Chunk document-type data as concatenated text."""
        lines = []
        for row in rows:
            cell_texts = [cell.strip() for cell in row if cell.strip()]
            if cell_texts:
                lines.append(" ".join(cell_texts))

        if not lines:
            return []

        full_text = f"[Sheet: {sheet_name}]\n" + "\n".join(lines)

        # Use simple chunking for document type
        chunk_size = 1000
        overlap = 200
        prefix = f"[Sheet: {sheet_name}]\n"

        if len(full_text) <= chunk_size:
            return [SheetChunk(text=full_text, sheet_name=sheet_name, chunk_type="document")]

        # Chunk the body text (without prefix), then re-add prefix
        body = "\n".join(lines)
        chunks = []
        start = 0

        while start < len(body):
            end = start + chunk_size - len(prefix)
            if end < len(body):
                # Try to break at newline
                last_nl = body[start:end].rfind("\n")
                if last_nl > (chunk_size - len(prefix)) // 2:
                    end = start + last_nl + 1

            chunk_body = body[start:end].strip()
            if chunk_body:
                chunks.append(SheetChunk(
                    text=prefix + chunk_body,
                    sheet_name=sheet_name,
                    chunk_type="document",
                ))

            start = end - overlap
            if start >= len(body):
                break

        return chunks

    def _decode(self, content: bytes) -> str:
        """Decode bytes to string with encoding detection."""
        for encoding in ["utf-8", "utf-8-sig", "cp949", "euc-kr", "utf-16", "latin-1"]:
            try:
                return content.decode(encoding)
            except (UnicodeDecodeError, UnicodeError):
                continue
        return content.decode("utf-8", errors="replace")
